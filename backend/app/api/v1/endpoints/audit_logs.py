"""
审计日志 API 端点
- GET /audit-logs   搜索审计日志（管理员）
- GET /audit-logs/export/excel  导出审计日志 Excel
"""

from fastapi import APIRouter, Depends
from fastapi.responses import Response
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_db_session, require_admin
from app.models.user import User
from app.schemas.common import ApiResponse, PaginatedData, PaginationParams
from app.schemas.notification import AuditLogOut, AuditLogSearchParams
from app.services.audit_service import AuditService

router = APIRouter()


@router.get("", response_model=ApiResponse[PaginatedData[AuditLogOut]], summary="搜索审计日志")
async def list_audit_logs(
    pagination: PaginationParams = Depends(),
    user_id: int = None,
    action: str = None,
    target_type: str = None,
    is_success: bool = None,
    _: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db_session),
):
    """管理员查看审计日志（不可修改）"""
    svc = AuditService(db)
    items, total = await svc.search(
        user_id=user_id,
        action=action,
        target_type=target_type,
        is_success=is_success,
        offset=pagination.offset,
        limit=pagination.page_size,
    )
    return ApiResponse.ok(
        data=PaginatedData.create(
            items=[AuditLogOut.model_validate(log) for log in items],
            total=total,
            page=pagination.page,
            page_size=pagination.page_size,
        )
    )


@router.get("/export/excel", summary="导出审计日志 Excel")
async def export_audit_logs_excel(
    user_id: int = None,
    action: str = None,
    _: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db_session),
):
    """将审计日志导出为 Excel（管理员）"""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    svc = AuditService(db)
    items, _ = await svc.search(user_id=user_id, action=action, limit=50000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "审计日志"

    headers = ["ID", "时间", "用户", "IP", "操作码", "操作描述", "对象类型", "对象ID", "是否成功"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    for row_idx, log in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=log.id)
        ws.cell(row=row_idx, column=2, value=log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "")
        ws.cell(row=row_idx, column=3, value=log.username or "")
        ws.cell(row=row_idx, column=4, value=log.ip_address or "")
        ws.cell(row=row_idx, column=5, value=log.action)
        ws.cell(row=row_idx, column=6, value=log.action_desc or "")
        ws.cell(row=row_idx, column=7, value=log.target_type or "")
        ws.cell(row=row_idx, column=8, value=log.target_id or "")
        ws.cell(row=row_idx, column=9, value="是" if log.is_success else "否")

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=audit_logs.xlsx"},
    )
