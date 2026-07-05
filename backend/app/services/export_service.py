"""
导出服务
支持将报销申请列表导出为 Excel / CSV / PDF 摘要。
"""

import csv
import io
from datetime import datetime
from typing import Sequence

from sqlalchemy.ext.asyncio import AsyncSession

from app.models.reimbursement import Reimbursement


class ExportService:
    """报销数据导出服务"""

    def __init__(self, db: AsyncSession) -> None:
        self._db = db

    async def export_excel(self, reimbursements: Sequence[Reimbursement]) -> bytes:
        """
        将报销申请列表导出为 Excel（.xlsx）。
        返回文件字节内容，由路由层写入 Response。
        """
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "报销申请列表"

        # 表头
        headers = [
            "申请ID", "申请人", "部门", "项目", "标题", "报销事由",
            "申报金额（元）", "OCR总金额（元）", "状态", "提交时间", "最后更新",
        ]
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        # 状态中文映射
        status_map = {
            "draft": "草稿",
            "submitted": "已提交",
            "reviewing": "审批中",
            "approved": "已通过",
            "rejected": "已驳回",
            "ready": "待线下报销",
        }

        # 数据行
        for row_idx, r in enumerate(reimbursements, start=2):
            applicant_name = r.applicant.full_name or r.applicant.username if r.applicant else ""
            department = r.applicant.department or "" if r.applicant else ""
            project_name = r.project.name if r.project else ""

            ws.cell(row=row_idx, column=1, value=r.id)
            ws.cell(row=row_idx, column=2, value=applicant_name)
            ws.cell(row=row_idx, column=3, value=department)
            ws.cell(row=row_idx, column=4, value=project_name)
            ws.cell(row=row_idx, column=5, value=r.title)
            ws.cell(row=row_idx, column=6, value=r.purpose or "")
            ws.cell(row=row_idx, column=7, value=r.declared_amount)
            ws.cell(row=row_idx, column=8, value=r.ocr_total_amount)
            ws.cell(row=row_idx, column=9, value=status_map.get(r.status.value, r.status.value))
            ws.cell(row=row_idx, column=10, value=r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "")
            ws.cell(row=row_idx, column=11, value=r.updated_at.strftime("%Y-%m-%d %H:%M") if r.updated_at else "")

        # 写入内存
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def export_csv(self, reimbursements: Sequence[Reimbursement]) -> bytes:
        """将报销申请列表导出为 CSV（UTF-8 BOM，兼容 Excel 直接打开）"""
        buf = io.StringIO()
        writer = csv.writer(buf)

        writer.writerow([
            "申请ID", "申请人", "部门", "项目", "标题", "报销事由",
            "申报金额（元）", "OCR总金额（元）", "状态", "提交时间",
        ])

        status_map = {
            "draft": "草稿", "submitted": "已提交", "reviewing": "审批中",
            "approved": "已通过", "rejected": "已驳回", "ready": "待线下报销",
        }

        for r in reimbursements:
            applicant_name = r.applicant.full_name or r.applicant.username if r.applicant else ""
            writer.writerow([
                r.id,
                applicant_name,
                r.applicant.department or "" if r.applicant else "",
                r.project.name if r.project else "",
                r.title,
                r.purpose or "",
                r.declared_amount,
                r.ocr_total_amount,
                status_map.get(r.status.value, r.status.value),
                r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
            ])

        # 添加 UTF-8 BOM 头，兼容 Windows Excel 直接打开
        return ("﻿" + buf.getvalue()).encode("utf-8")

    async def export_pdf_summary(self, reimbursements: Sequence[Reimbursement]) -> bytes:
        """
        将报销申请列表导出为 PDF 摘要报告。
        使用 reportlab 生成简单的表格式 PDF。
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=1 * cm, leftMargin=1 * cm)
        styles = getSampleStyleSheet()
        story = []

        # 标题
        story.append(Paragraph("报销申请汇总报告", styles["Title"]))
        story.append(Paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
        story.append(Spacer(1, 0.5 * cm))

        # 表格数据
        data = [["ID", "申请人", "项目", "标题", "申报金额", "状态", "日期"]]
        status_map = {
            "draft": "草稿", "submitted": "已提交", "reviewing": "审批中",
            "approved": "已通过", "rejected": "已驳回", "ready": "待线下报销",
        }
        for r in reimbursements:
            data.append([
                str(r.id),
                r.applicant.full_name or r.applicant.username if r.applicant else "",
                r.project.name[:12] if r.project else "",
                r.title[:20],
                f"¥{r.declared_amount:.2f}",
                status_map.get(r.status.value, r.status.value),
                r.created_at.strftime("%Y-%m-%d") if r.created_at else "",
            ])

        table = Table(data, colWidths=[1.5 * cm, 3 * cm, 4 * cm, 6 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(table)
        doc.build(story)
        return buf.getvalue()
